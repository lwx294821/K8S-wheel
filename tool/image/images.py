#!/usr/bin/python3
# -*- coding:utf-8 -*-
# Author:Lishuwen
# @Time:2019/5/20 17:26
import argparse
import os
import subprocess

import openpyxl
import yaml

'''
 基于ansible生成的镜像清单生成所有镜像的Dockerfile，为在DockerHub上下载镜像做准备 
'''

DOCKERFILE = './download_containers/'
DOCKERHUB = 'docker.io'
ACCOUNT = 'd20190308/k8s-images'


def reqargs():
    parser = argparse.ArgumentParser(description='Tool for Auto Manager Images and File.')
    required = parser.add_argument_group('required arguments')
    optional = parser.add_argument_group('optional arguments')
    # 必选参数
    required.add_argument('--case-controller', '-c', dest='case', type=int,
                          help='''Controller Step.''',
                          required=False)

    # 可选参数，并且设置了默认值，需要结合当前环境确保对象存在
    optional.add_argument('--images-inventory', '-i', dest='images_inventory', type=str,
                          help='Download Needed_Images Inventory YAML File',
                          default='./k8s-images.yaml')
    optional.add_argument('--files-inventory', '-f', dest='files_inventory', type=str,
                          help='Download Needed_File Inventory YAML File', default='./k8s-files.yaml')
    optional.add_argument('--report-containers', '-a', dest='report_containers', type=str,
                          help='Excel Containers Save Path',
                          default='./container.xlsx')
    optional.add_argument('--report-files', '-b', dest='report_files', type=str, help='Excel Files Save Path',
                          default='./files.xlsx')

    # 返回获取的命令行参数
    return parser.parse_args()


def process():
    args = reqargs()
    case = args.case

    if case == -1:
        print('''0 - create dockerfile. "python3 images.py -c 0 -i ./2019-05-22-images-inventory.yaml"
1 - Generate Report About  Depend Containers. "python3 images.py -c 1 -i ./2019-05-22-images-inventory.yaml"
2 - Generate Report About Depend Files."python3 images.py -c 2 -f ./2019-05-22-files-inventory.yaml"
3 - Pull Images from Dockerhub. "python3 images.py -c 3 -i ./2019-05-22-images-inventory.yaml"''', )
    elif case == 0:
        if not args.images_inventory:
            print('Please appoint one special image yaml')
            return
        mkdir(args.images_inventory)
    elif case == 1:
        if not args.images_inventory:
            print('Please appoint one special image yaml')
            return
        reportcontainers(args.images_inventory, args.report_containers)
    elif case == 2:
        if not args.files_inventory:
            print('Please appoint one special files yaml')
            return
        reportfiles(args.files_inventory, args.report_files)
    elif case == 3:
        if args.images_inventory:
            pull(args.images_inventory)
    else:
        print('Input at least one argument or argument error.')


def reportcontainers(images, report):
    global i
    i = 1
    with open(images, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        outwb = openpyxl.Workbook()
        outws = outwb.create_sheet(index=0)
        for k in res.keys():
            outws.cell(i, 1).value = k
            outws.cell(i, 2).value = res.get(k)['repo']
            outws.cell(i, 3).value = res.get(k)['tag']
            outws.cell(i, 4).value = "FROM %s:%s" % (res.get(k)['repo'], str(res.get(k)['tag']))
            i = i + 1
        outwb.save(report)


def reportfiles(files, report):
    global i
    i = 1
    with open(files, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        outwb = openpyxl.Workbook()
        outws = outwb.create_sheet(index=0)
        for k in res.keys():
            outws.cell(i, 1).value = k
            outws.cell(i, 2).value = res.get(k)['dest']
            outws.cell(i, 3).value = res.get(k)['url']

            i = i + 1
        outwb.save(report)


def pull(images):
    with open(images, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        for r in res.keys():
            if isinstance(r, str):
                repo = res.get(r)['repo']
                tag = res.get(r)['tag']
                if isinstance(repo, str) and repo.startswith(DOCKERHUB):
                    build({'from': DOCKERFILE + repo + '/Dockerfile', 'tag': r + ':' + str(tag)})
                else:
                    old_tag = ACCOUNT + ":" + r
                    new_tag = repo.split('/')[-1] + ":" + str(tag)
                    p = 'docker pull %s' % old_tag
                    t = 'docker tag %s %s' % (old_tag, new_tag)
                    r = 'docker rmi %s' % old_tag
                    pout = subprocess.Popen(p, shell=True)
                    for line in pout:
                        print(line.strip().decode('utf-8'))

                        tout = subprocess.Popen(t, shell=True)
                    for line in tout:
                        print(line.strip().decode('utf-8'))

                    rout = subprocess.Popen(r, shell=True)
                    for line in rout:
                        print(line.strip().decode('utf-8'))


'''

创建Dockerfile

'''


def mkdir(image_invent):
    del_file(DOCKERFILE)
    with open(image_invent, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        for r in res:
            if isinstance(r, str):
                dockerfile(res.get(r)['repo'], res.get(r)['tag'])


def check():
    p = subprocess.Popen("docker images -q | xargs docker inspect -f '{{.Id}}'", shell=True, stdout=subprocess.PIPE)
    out = p.stdout.readlines()
    for line in out:
        sha256 = line.strip().decode('utf-8').split(":")[1]
        print(sha256)


def dockerfile(sub, tag):
    path = DOCKERFILE + '/' + sub
    doc = path + '/Dockerfile'
    if not os.path.exists(path):
        os.makedirs(path)
    with open(doc, "w", encoding="UTF-8") as f:
        f.write('FROM ' + sub + ':' + str(tag))


'''
docker build -f Dockerfile .
'''


def build(d):
    if isinstance(d, dict):
        b = 'docker build -f %s  -t %s .' % (d.get('from'), d.get('tag'))
        print(b)
        # path = d[0:d.rfind('/')]
        # p = subprocess.Popen(b, shell=True, stdout=subprocess.PIPE)
        # out = p.stdout.readlines()
        # for line in out:
        #     print(line.strip().decode('utf-8'))


def del_file(path):
    if not os.path.exists(path):
        os.makedirs(path)
        return
    ls = os.listdir(path)
    for i in ls:
        c_path = os.path.join(path, i)
        if os.path.isdir(c_path):
            del_file(c_path)
        else:
            os.remove(c_path)


if __name__ == "__main__":
    process()
