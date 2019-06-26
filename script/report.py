#!/usr/bin/python3
# -*- coding:utf-8 -*-
# Author:Lishuwen
# @Time:2019/5/20 17:26
import argparse
import os
import subprocess

import openpyxl
import yaml

'''
 基于ansible生成的镜像清单生成所有镜像的Dockerfile，为在DockerHub上下载镜像做准备 
'''

DOCKERFILE = './download_containers/'
DOCKERHUB = 'docker.io'
ACCOUNT = 'd20190308/k8s-images'


def reqargs():
    parser = argparse.ArgumentParser(description='Tool for Auto Manager Images and File.')
    required = parser.add_argument_group('required arguments')
    optional = parser.add_argument_group('optional arguments')
    # 必选参数
    required.add_argument('--case-controller', '-c', dest='case', type=int,
                          help='''Controller Step.''',
                          required=False)

    # 可选参数，并且设置了默认值，需要结合当前环境确保对象存在
    optional.add_argument('--images-inventory', '-i', dest='images_inventory', type=str,
                          help='Download Needed_Images Inventory YAML File',
                          default='./k8s-images.yaml')
    optional.add_argument('--files-inventory', '-f', dest='files_inventory', type=str,
                          help='Download Needed_File Inventory YAML File', default='./k8s-files.yaml')
    optional.add_argument('--report-containers', '-a', dest='report_containers', type=str,
                          help='Excel Containers Save Path',
                          default='./container.xlsx')
    optional.add_argument('--report-files', '-b', dest='report_files', type=str, help='Excel Files Save Path',
                          default='./files.xlsx')

    # 返回获取的命令行参数
    return parser.parse_args()


def process():
    pass


def reportcontainers(images, report):
    global i
    i = 1
    with open(images, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        outwb = openpyxl.Workbook()
        outws = outwb.create_sheet(index=0)
        for k in res.keys():
            outws.cell(i, 1).value = k
            outws.cell(i, 2).value = res.get(k)['repo']
            outws.cell(i, 3).value = res.get(k)['tag']
            outws.cell(i, 4).value = "FROM %s:%s" % (res.get(k)['repo'], str(res.get(k)['tag']))
            i = i + 1
        outwb.save(report)


def reportfiles(files, report):
    global i
    i = 1
    with open(files, mode='r', encoding='UTF_8') as f:
        res = yaml.load(f, Loader=yaml.SafeLoader)
        outwb = openpyxl.Workbook()
        outws = outwb.create_sheet(index=0)
        for k in res.keys():
            outws.cell(i, 1).value = k
            outws.cell(i, 2).value = res.get(k)['dest']
            outws.cell(i, 3).value = res.get(k)['url']

            i = i + 1
        outwb.save(report)


def check():
    p = subprocess.Popen("docker images -q | xargs docker inspect -f '{{.Id}}'", shell=True, stdout=subprocess.PIPE)
    out = p.stdout.readlines()
    for line in out:
        sha256 = line.strip().decode('utf-8').split(":")[1]
        print(sha256)


def del_file(path):
    if not os.path.exists(path):
        os.makedirs(path)
        return
    ls = os.listdir(path)
    for i in ls:
        c_path = os.path.join(path, i)
        if os.path.isdir(c_path):
            del_file(c_path)
        else:
            os.remove(c_path)


if __name__ == "__main__":
    process()
